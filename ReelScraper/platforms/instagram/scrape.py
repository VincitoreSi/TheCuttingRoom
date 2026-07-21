#!/usr/bin/env python3
"""
scrape_reels.py — portable Instagram Reels scraper (SAFE guest mode)

Scrapes up to N reels per creator with FULL data (metrics + CDN links + audio),
plus each creator's follower count / profile meta (needed for virality analysis),
and writes an .xlsx (+ raw .json backup + profiles_meta.json).

SAFETY:
  * Guest mode only — never sends a logged-in session (sessionid). No account is
    attached to any request, so your account cannot be banned.
  * Randomized human-like pacing + a circuit breaker that STOPS if Instagram
    rate-limits the IP (avoids getting the IP blocked).
  * Resume — creators already present in any reels_raw*.json are skipped.

USAGE (run from inside this folder):
  python scrape_reels.py nasa instagram            # scrape specific handles
  python scrape_reels.py --file pages.txt           # scrape handles from a file (.txt or .xlsx)
  python scrape_reels.py --file list.xlsx --limit 100
  # parallel (faster) — run two in separate terminals, then merge:
  python scrape_reels.py --file pages.txt --worker 0 --workers 2
  python scrape_reels.py --file pages.txt --worker 1 --workers 2
  python merge_reels.py
  # then analyze:
  python analyze_virality.py

OPTIONS:
  --file PATH     input file: .txt (one handle/URL per line) or .xlsx (handles auto-detected)
  --limit N       max reels per creator (default: niche_config.reels_per_creator, else 250)
  --out NAME      output xlsx name (default Reels_Data.xlsx)
  --worker i      this worker's index (for parallel runs)
  --workers N     total number of workers (default 1)

Requires: pip install openpyxl
"""
import sys, json, time, re, random, argparse, logging
from http.cookiejar import CookieJar
from pathlib import Path
import urllib.request, urllib.parse, urllib.error
import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from core.logsetup import setup_logging  # noqa: E402
from core.atomicio import write_text_atomic, atomic_path  # noqa: E402
from core.stopflag import install_stop_handler, stop_requested, sleep_unless_stopped  # noqa: E402

HERE = Path(__file__).parent
log = logging.getLogger("instagram.scrape")

# ---- constants ----
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36"
APP = "936619743392459"   # public web app id (constant)
PAGE_SIZE = 12
PAGE_DELAY = (4.0, 8.0)        # seconds between pages (randomized)
CREATOR_DELAY = (10.0, 20.0)   # seconds between creators (randomized)
MAX_429_IN_A_ROW = 3           # circuit breaker

# ---- guest session state ----
_cookie = ""; _csrf = ""; _req_count = 0; _consec_429 = 0

def new_guest_session():
    """Fetch a fresh GUEST csrftoken/mid (NO login, NO sessionid)."""
    global _cookie, _csrf
    cj = CookieJar()
    op = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
    op.addheaders = [("User-Agent", UA)]
    op.open("https://www.instagram.com/", timeout=30).read()
    jar = {c.name: c.value for c in cj}
    assert "sessionid" not in jar, "guest session unexpectedly logged in!"
    _csrf = jar.get("csrftoken", "")
    _cookie = "; ".join(f"{k}={v}" for k, v in jar.items())
    return bool(_csrf)

def _http(url, method="GET", data=None, referer=None, timeout=30):
    headers = {
        "User-Agent": UA, "X-IG-App-ID": APP, "Sec-Fetch-Site": "same-origin",
        "Cookie": _cookie, "X-CSRFToken": _csrf,
        "Content-Type": "application/x-www-form-urlencoded",
    }
    if referer:
        headers["Referer"] = referer
    body = data.encode() if isinstance(data, str) else data
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.status, r.read().decode("utf-8", "replace")

class RateLimited(Exception):
    pass

def _parse_count(s):
    """'131776' / '131,776' / '131.7K' / '12M' / '1.2B' -> int."""
    s = str(s).strip().replace(",", "")
    m = re.match(r"([\d.]+)\s*([KMB]?)", s, re.I)
    if not m:
        return None
    n = float(m.group(1))
    mult = {"K": 1e3, "M": 1e6, "B": 1e9}.get(m.group(2).upper(), 1)
    return int(n * mult)

def _page_scrape_profile(username):
    """Fallback: pull id + follower count straight from the profile page HTML (guest)."""
    try:
        _, t = _http(f"https://www.instagram.com/{username}/", referer="https://www.instagram.com/")
    except urllib.error.HTTPError:
        return None, None
    m = re.search(r'"profile_id":"(\d+)"', t) or re.search(r'"id":"(\d+)","is_private"', t)
    uid = m.group(1) if m else None
    # exact count if embedded, else the abbreviated count in the og:description meta
    fm = (re.search(r'"edge_followed_by":\{"count":(\d+)\}', t)
          or re.search(r'"follower_count":(\d+)', t))
    followers = int(fm.group(1)) if fm else None
    if followers is None:
        om = re.search(r'content="([\d.,]+[KMB]?)\s+Followers', t)
        if om:
            followers = _parse_count(om.group(1))
    return uid, followers

def _page_scrape_id(username):
    return _page_scrape_profile(username)[0]

def get_profile(username):
    """GUEST profile hydration -> {user_id, followers, full_name, ...}.

    Uses the public web_profile_info endpoint (needed for follower count, which
    drives engagement rate + reach multiplier). Falls back to a page-scrape for
    the id + followers if the API is unavailable. Still 100% guest — no sessionid."""
    global _consec_429
    url = "https://www.instagram.com/api/v1/users/web_profile_info/?" + urllib.parse.urlencode({"username": username})
    ref = f"https://www.instagram.com/{username}/"
    for attempt in range(4):
        try:
            _, t = _http(url, referer=ref)
            u = (json.loads(t).get("data") or {}).get("user") or {}
            _consec_429 = 0
            if not u:
                break
            return {
                "username": u.get("username") or username,
                "full_name": u.get("full_name", ""),
                "user_id": u.get("id"),
                "followers": (u.get("edge_followed_by") or {}).get("count"),
                "following": (u.get("edge_follow") or {}).get("count"),
                "posts": (u.get("edge_owner_to_timeline_media") or {}).get("count"),
                "category": u.get("category_name") or u.get("category") or "",
                "is_verified": u.get("is_verified"),
                "is_private": u.get("is_private"),
                "is_business": u.get("is_business_account"),
                "biography": u.get("biography", ""),
                "external_url": u.get("external_url", ""),
            }
        except urllib.error.HTTPError as e:
            if e.code == 429:
                _consec_429 += 1
                if _consec_429 >= MAX_429_IN_A_ROW:
                    raise RateLimited(f"{_consec_429} consecutive 429s — stopping to protect the IP")
                time.sleep(15 * (attempt + 1) + random.uniform(0, 5))
                try: new_guest_session()
                except Exception: pass
                continue
            if e.code in (401, 403):
                new_guest_session(); time.sleep(2); continue
            break
        except Exception:
            break
    # fallback: resolve id (and followers if the HTML exposes them) so we can still scrape
    uid, followers = _page_scrape_profile(username)
    if uid:
        return {"username": username, "user_id": uid, "full_name": "", "followers": followers}
    return None

def fetch_page(user_id, username, max_id):
    global _consec_429
    params = {"target_user_id": user_id, "page_size": str(PAGE_SIZE), "include_feed_video": "true"}
    if max_id:
        params["max_id"] = max_id
    body = urllib.parse.urlencode(params)
    ref = f"https://www.instagram.com/{username}/reels/"
    for attempt in range(4):
        try:
            _, t = _http("https://www.instagram.com/api/v1/clips/user/", "POST", body, ref)
            _consec_429 = 0
            return json.loads(t)
        except urllib.error.HTTPError as e:
            if e.code == 429:
                _consec_429 += 1
                if _consec_429 >= MAX_429_IN_A_ROW:
                    raise RateLimited(f"{_consec_429} consecutive 429s — stopping to protect the IP")
                w = 15 * (attempt + 1) + random.uniform(0, 5)
                log.warning("429 — backing off + refreshing guest session", extra={"wait_s": round(w), "consec_429": _consec_429})
                time.sleep(w)
                try: new_guest_session()
                except Exception: pass
                continue
            if e.code in (401, 403):
                log.info("refreshing guest session", extra={"code": e.code})
                new_guest_session(); time.sleep(2); continue
            raise
    raise RuntimeError("page failed after retries")

def scrape_creator(username, limit):
    global _req_count
    prof = get_profile(username)
    if not prof or not prof.get("user_id"):
        log.warning("could not resolve creator (private/typo/blocked)", extra={"creator": username})
        return None, []
    uid = prof["user_id"]
    fol = prof.get("followers")
    log.info("resolved", extra={"creator": username, "followers": fol})
    items, max_id, page = [], None, 0
    while len(items) < limit:
        page += 1
        _req_count += 1
        if _req_count % 40 == 0:
            new_guest_session()
        j = fetch_page(uid, username, max_id)
        batch = [x.get("media", x) for x in j.get("items", [])]
        items.extend(batch)
        log.info("page", extra={"creator": username, "page": page, "added": len(batch), "total": len(items)})
        pg = j.get("paging_info") or {}
        max_id = pg.get("max_id") if pg.get("more_available") else None
        if not max_id:
            break
        time.sleep(random.uniform(*PAGE_DELAY))
    return prof, items[:limit]

def flatten(m, creator, followers=None):
    imgs = (m.get("image_versions2") or {}).get("candidates") or []
    vids = m.get("video_versions") or []
    cm = m.get("clips_metadata") or {}
    music = (cm.get("music_info") or {}).get("music_asset_info") or cm.get("original_sound_info") or {}
    ta = m.get("taken_at")
    return {
        "creator": creator,
        "creator_followers": followers,
        "shortcode": m.get("code"),
        "url": f"https://www.instagram.com/reel/{m.get('code')}/",
        "id": m.get("id"), "pk": m.get("pk"),
        "taken_at": ta,
        "taken_at_iso": time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime(ta)) if ta else "",
        "media_type": m.get("media_type"), "product_type": m.get("product_type"),
        "play_count": m.get("play_count") or m.get("view_count"),
        "ig_play_count": m.get("ig_play_count"),
        "like_count": m.get("like_count"), "comment_count": m.get("comment_count"),
        "reshare_count": m.get("reshare_count"), "save_count": m.get("save_count"),
        "video_duration": m.get("video_duration"),
        "width": m.get("original_width"), "height": m.get("original_height"),
        "has_audio": m.get("has_audio"),
        "username": (m.get("user") or {}).get("username"),
        "full_name": (m.get("user") or {}).get("full_name"),
        "user_id": (m.get("user") or {}).get("pk") or (m.get("user") or {}).get("id"),
        "music_title": music.get("title") or music.get("song_name") or music.get("original_audio_title"),
        "music_artist": music.get("display_artist") or music.get("artist_name"),
        "music_audio_url": music.get("progressive_download_url") or music.get("fast_start_progressive_download_url"),
        "video_url_best": vids[0]["url"] if vids else "",
        "video_urls_all": " | ".join(v["url"] for v in vids),
        "thumbnail_best": imgs[0]["url"] if imgs else "",
        "thumbnail_urls_all": " | ".join(f"{i.get('width')}x{i.get('height')}:{i['url']}" for i in imgs),
        "caption": (m.get("caption") or {}).get("text", "") if m.get("caption") else "",
    }

# ---- input parsing ----
def norm_handle(s):
    s = str(s).strip()
    m = re.search(r"instagram\.com/([A-Za-z0-9_.]+)", s)
    if m: return m.group(1)
    s = s.lstrip("@").strip("/")
    return s if re.fullmatch(r"[A-Za-z0-9_.]{1,40}", s) else None

def read_input_file(path):
    p = Path(path)
    handles = []
    # A fresh install has no pages.txt yet (only pages.txt.example ships). Treat a
    # missing file as "no handles" so main() reaches its clear "no creators given —
    # fill pages.txt" guidance and exits 1, instead of dying on an uncaught
    # FileNotFoundError traceback that tells a new user nothing.
    if not p.exists():
        return handles
    if p.suffix.lower() in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(p, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, str):
                        h = None
                        m = re.search(r"instagram\.com/([A-Za-z0-9_.]+)", cell)
                        if m: h = m.group(1)
                        if h and h not in handles:
                            handles.append(h)
    else:
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            h = norm_handle(line)
            if h and h not in handles:
                handles.append(h)
    return handles

def rows_and_summary(raw_all, meta_all):
    """Flat xlsx rows + the per-creator tally, rebuilt from the FULL raw set.

    `save_outputs` rewrites both files wholesale, so what it is handed has to describe
    everything on disk — not just the creators this run happened to touch. Deriving them
    here rather than accumulating during the loop is what makes a resumed run safe:
    skipped creators are still in `raw_all`, so they still make it into both outputs.
    `flatten` is pure, so re-deriving a previously-saved creator's rows is free of
    side effects and produces exactly what the earlier run wrote.
    """
    rows, summary = [], []
    for c, items in raw_all.items():
        fol = (meta_all.get(c) or {}).get("followers")
        for m in items:
            rows.append(flatten(m, c, fol))
        summary.append((c, len(items)))
    return rows, summary


def save_outputs(out_xlsx, raw_json, all_rows, raw_all, summary):
    """Rewrite the corpus + workbook wholesale, atomically.

    Both writes go through a temp file and a rename. `reels_raw.json` IS the corpus, and
    the old `write_text` truncated it at open(): a signal landing one instruction later
    left it short, the next run's `except: raw_all = {}` read that as an empty corpus, and
    the run after that wrote `{}` over everything that had survived — silently, rc 0. The
    Stop button makes signals here an ordinary event, so this is no longer a rare window.
    """
    write_text_atomic(raw_json, json.dumps(raw_all, ensure_ascii=False, indent=1))
    out = openpyxl.Workbook()
    sh = out.active; sh.title = "Reels"
    cols = list(all_rows[0].keys()) if all_rows else []
    if cols:
        sh.append(cols)
        for r in all_rows:
            sh.append([r.get(c) for c in cols])
    s2 = out.create_sheet("Summary")
    s2.append(["creator", "reels_scraped"])
    for c, n in summary:
        s2.append([c, n])
    s2.append(["TOTAL", sum(n for _, n in summary)])
    # openpyxl opens a ZipFile on the path it is handed and streams the whole workbook into
    # it, so the DESTINATION is truncated for the entire serialization — a far wider window
    # than a json dump. Save to the temp, then promote.
    with atomic_path(out_xlsx) as tmp_xlsx:
        out.save(tmp_xlsx)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("handles", nargs="*", help="instagram handles or profile URLs")
    ap.add_argument("--file", help="input file (.txt or .xlsx)")
    ap.add_argument("--limit", type=int, default=None,
                    help="max reels per creator (default: niche_config.reels_per_creator, else 250)")
    ap.add_argument("--out", default="Reels_Data.xlsx")
    ap.add_argument("--worker", type=int, default=0)
    ap.add_argument("--workers", type=int, default=1)
    args = ap.parse_args()
    setup_logging("scrape", platform="instagram")
    # Take SIGTERM over before anything is written. The hub's Stop button signals this
    # process group; under the default disposition that would end the run between two
    # bytecodes, which is how reels_raw.json used to end up truncated. With a handler
    # installed the signal only sets a flag, and the loop below stops at the next creator
    # boundary — where everything is already on disk.
    install_stop_handler()

    # reel limit: CLI wins, else niche_config.json, else 250
    if args.limit is None:
        args.limit = 250
        try:
            cfg = json.loads((HERE / "niche_config.json").read_text(encoding="utf-8"))
            args.limit = int(cfg.get("reels_per_creator") or 250)
        except Exception:
            pass

    # gather creators
    creators = []
    for h in args.handles:
        nh = norm_handle(h)
        if nh and nh not in creators:
            creators.append(nh)
    if args.file:
        for h in read_input_file(args.file):
            if h not in creators:
                creators.append(h)
    if not creators and (HERE / "pages.txt").exists():
        creators = read_input_file(HERE / "pages.txt")
    if not creators:
        log.error("no creators given — pass handles, --file, or fill pages.txt")
        sys.exit(1)

    suffix = f"_w{args.worker}" if args.workers > 1 else ""
    out_xlsx = HERE / (args.out if not suffix else f"{Path(args.out).stem}{suffix}.xlsx")
    raw_json = HERE / f"reels_raw{suffix}.json"
    meta_json = HERE / f"profiles_meta{suffix}.json"

    if not new_guest_session():
        log.error("could not establish guest session"); sys.exit(1)
    log.info("guest session OK (no sessionid attached)")

    mine = creators[args.worker::args.workers]
    # resume: skip creators already saved in any reels_raw*.json
    done = set()
    for f in HERE.glob("reels_raw*.json"):
        try: done.update(json.loads(f.read_text(encoding="utf-8")).keys())
        except Exception: pass
    todo = [c for c in mine if c not in done]
    log.info("plan", extra={"worker": args.worker, "workers": args.workers, "assigned": len(mine),
                            "already_saved": len(mine) - len(todo), "scraping": len(todo)})

    # profile meta (followers etc.) — preserve any already saved, add as we go
    meta_all = {}
    if meta_json.exists():
        try: meta_all = json.loads(meta_json.read_text(encoding="utf-8"))
        except Exception: meta_all = {}

    # Seed from what is already on disk, exactly as meta_all above — and for the same
    # reason, which was missed here. `save_outputs` writes this dict WHOLESALE, so every
    # creator absent from it is deleted from the file. `todo` deliberately excludes
    # creators already saved, so starting empty meant a resumed run rewrote the corpus
    # down to just this run's additions; and when nothing was left to do it wrote `{}`
    # over everything and still logged DONE with rc 0. Adding one handle to a watchlist
    # of five destroyed the other four; running the pipeline twice destroyed the lot and
    # left analyze reporting "no scraped data — scrape first".
    raw_all = {}
    if raw_json.exists():
        try:
            raw_all = json.loads(raw_json.read_text(encoding="utf-8")) or {}
        except Exception:
            raw_all = {}
    if not isinstance(raw_all, dict):
        raw_all = {}
    kept = len(raw_all)
    stopped = False
    by_request = False
    for n, c in enumerate(todo, 1):
        # The one point at which stopping is free: everything up to here is already saved.
        # Reuses the same `stopped` flag the rate-limit circuit breaker sets, so both early
        # exits share the one final-save path below.
        if stop_requested():
            log.warning("stop requested — ending after the last saved creator",
                        extra={"scraped_this_run": n - 1, "remaining": len(todo) - n + 1})
            stopped = by_request = True
            break
        log.info("creator start", extra={"i": n, "of": len(todo), "creator": c})
        try:
            prof, items = scrape_creator(c, args.limit)
        except RateLimited as e:
            log.error("CIRCUIT BREAKER — saving partial progress and exiting", extra={"reason": str(e)})
            stopped = True
            break
        if prof:
            meta_all[c] = prof
        raw_all[c] = items
        all_rows, summary = rows_and_summary(raw_all, meta_all)
        # META BEFORE CORPUS, and that order is load-bearing. Resume keys on presence in
        # raw_all, so dying between the two writes with the corpus written FIRST left the
        # creator present in the corpus and absent from profiles_meta.json — skipped
        # forever on every later run, its follower count never re-fetched, and
        # core/virality.py divides engagement_rate and reach_multiplier by followers. That
        # creator's two strongest signals would be permanently null. This way round the
        # worst case is a meta entry for a creator not yet in the corpus, which the next
        # run simply overwrites when it scrapes them.
        write_text_atomic(meta_json, json.dumps(meta_all, ensure_ascii=False, indent=1))
        save_outputs(out_xlsx, raw_json, all_rows, raw_all, summary)
        # Interruptible: PEP 475 RESUMES a bare time.sleep after the stop handler returns,
        # so the flag alone could not shorten a 10-20s wait and a Stop press would look
        # ignored for up to twenty seconds.
        sleep_unless_stopped(random.uniform(*CREATOR_DELAY))

    all_rows, summary = rows_and_summary(raw_all, meta_all)
    write_text_atomic(meta_json, json.dumps(meta_all, ensure_ascii=False, indent=1))
    save_outputs(out_xlsx, raw_json, all_rows, raw_all, summary)
    # A stop is a normal outcome, not a failure — the process still exits 0 below, and the
    # hub tells a stop from a crash by its own marker, never by the return code.
    status = ("STOPPED (by request)" if by_request
              else "STOPPED EARLY (rate limit)" if stopped else "DONE")
    # Report the corpus TOTAL and what this run added separately. The old line reported
    # only this run's additions, so a resumed run that legitimately had nothing to do
    # said "0 reels across 0 creators" — indistinguishable from a scrape that failed to
    # pull anything, and the phrasing the hub then showed on the board.
    added = len(raw_all) - kept
    log.info("%s — %d reels across %d creators (+%d new) -> %s",
             status, len(all_rows), len(summary), added, out_xlsx.name,
             extra={"status": status, "reels": len(all_rows), "creators": len(summary),
                    "new_creators": added, "per_creator": dict(summary)})
    if args.workers > 1:
        log.info("run `python merge.py` after all workers finish to combine into one xlsx")

if __name__ == "__main__":
    main()
