#!/usr/bin/env python3
"""
core/virality.py — shared, platform-agnostic virality engine.

Consumes normalized core.schema.Content records (from any platform's normalize.py)
and computes four signals + a blended 0-100 score + tier. Identical math for
Instagram, X, and YouTube — each platform just tunes weights/tiers in its own
niche_config.json.

  engagement_rate  = (likes+comments+shares+saves) / followers
  reach_multiplier = plays / followers
  outlier_score    = plays / that creator's MEDIAN plays
  velocity         = plays / days since posting

Each signal is percentile-normalized across the dataset, then blended by weights.
"""
import io, json, csv, time, bisect
from pathlib import Path
from statistics import median
import openpyxl

from core.atomicio import write_bytes_atomic, atomic_path

from core.schema import Content

DEFAULT_WEIGHTS = {"reach_multiplier": 0.35, "outlier_score": 0.25,
                   "engagement_rate": 0.25, "velocity": 0.15}
DEFAULT_TIERS = [{"label": "Viral", "min_score": 85}, {"label": "High", "min_score": 70},
                 {"label": "Above Average", "min_score": 50}, {"label": "Normal", "min_score": 0}]
METRIC_KEYS = ["engagement_rate", "reach_multiplier", "outlier_score", "velocity"]


def load_config(path):
    cfg = {}
    try:
        cfg = json.loads(Path(path).read_text(encoding="utf-8"))
    except Exception:
        pass
    v = cfg.get("virality") or {}
    weights = {k: float(w) for k, w in (v.get("weights") or DEFAULT_WEIGHTS).items()
               if k in METRIC_KEYS and float(w) > 0} or dict(DEFAULT_WEIGHTS)
    tiers = sorted(v.get("tiers") or DEFAULT_TIERS, key=lambda t: -t.get("min_score", 0))
    return weights, tiers, int(v.get("top_n") or 100)


def tier_threshold(tiers, label):
    """The min_score a `label` demands, read from a tiers list (as `load_config` returns it),
    or None if the label is not one of the configured tiers. Keeping this here means the media
    gate and the scoring engine read the SAME labels/thresholds — they cannot drift apart."""
    for t in tiers or []:
        if t.get("label") == label:
            return t.get("min_score", 0)
    return None


def resolve_media_filter(path):
    """Resolve `virality.media_filter` in a platform's niche_config into an effective
    (min_score, max_downloads) the media stage can act on.

    `min_tier` (a label) maps to a score threshold via the SAME tiers `load_config` returns,
    so a user picking "Viral" gets exactly the scoring engine's Viral cutoff. An explicit
    `min_score` overrides the tier-derived one. No `media_filter` (or an unknown tier label)
    resolves to no gate — (None, None) — so nothing changes until a user opts in."""
    cfg = {}
    try:
        cfg = json.loads(Path(path).read_text(encoding="utf-8"))
    except Exception:
        pass
    mf = ((cfg.get("virality") or {}).get("media_filter")) or {}
    _, tiers, _ = load_config(path)
    min_score = None
    if mf.get("min_tier"):
        min_score = tier_threshold(tiers, mf["min_tier"])
    if mf.get("min_score") is not None:
        min_score = mf["min_score"]
    max_downloads = mf.get("max_downloads")
    return min_score, max_downloads


def _pct_ranker(values):
    s = sorted(v for v in values if v is not None)
    n = len(s)
    def rank(v):
        return None if (v is None or n == 0) else bisect.bisect_right(s, v) / n
    return rank


def _tier(score, tiers):
    if score is None:
        return "n/a"
    for t in tiers:
        if score >= t.get("min_score", 0):
            return t.get("label", "")
    return tiers[-1].get("label", "") if tiers else ""


def analyze(contents, weights=None, tiers=None, now=None):
    """contents: list[Content] -> list[dict] rows with signals + virality_score + tier."""
    weights = weights or dict(DEFAULT_WEIGHTS)
    tiers = tiers or list(DEFAULT_TIERS)
    now = now or time.time()

    # per-creator median plays (over items with plays > 0)
    by_creator = {}
    for c in contents:
        by_creator.setdefault(c.creator, []).append(c)
    creator_median = {}
    for cr, items in by_creator.items():
        pl = [c.plays for c in items if c.plays]
        creator_median[cr] = median(pl) if pl else None

    rows = []
    for c in contents:
        f = c.creator_followers
        plays = c.plays
        eng = c.engagements
        med = creator_median.get(c.creator)
        age_days = max((now - c.posted_ts) / 86400.0, 0.5) if c.posted_ts else None
        row = c.to_row()
        row.update({
            "age_days": round(age_days, 2) if age_days is not None else None,
            "engagement_rate": (eng / f) if f else None,
            "reach_multiplier": (plays / f) if (f and plays) else None,
            "outlier_score": (plays / med) if (plays and med) else None,
            "velocity": (plays / age_days) if (plays and age_days) else None,
            "_content": c,
        })
        rows.append(row)

    rankers = {k: _pct_ranker([r.get(k) for r in rows]) for k in METRIC_KEYS}
    wtotal = sum(weights.get(k, 0) for k in METRIC_KEYS)
    for r in rows:
        num = wsum = 0.0
        for k in METRIC_KEYS:
            w = weights.get(k, 0)
            if w <= 0:
                continue
            p = rankers[k](r.get(k))
            if p is not None:
                num += w * p; wsum += w
        r["virality_score"] = round(100 * num / wsum, 1) if wsum > 0 else None
        r["coverage"] = round(wsum / wtotal, 2) if wtotal else 0
        r["tier"] = _tier(r["virality_score"], tiers)
    rows.sort(key=lambda r: (r["virality_score"] is None, -(r["virality_score"] or 0)))
    return rows


# ---- outputs ----
REEL_COLS = [
    ("platform", "platform"), ("creator", "creator"), ("creator_followers", "followers"),
    ("url", "url"), ("posted_iso", "posted"), ("age_days", "age_days"),
    ("plays", "plays"), ("likes", "likes"), ("comments", "comments"),
    ("shares", "shares"), ("saves", "saves"), ("engagements", "engagements"),
    ("duration_s", "duration_s"),
    ("engagement_rate", "engagement_rate"), ("reach_multiplier", "reach_multiplier"),
    ("outlier_score", "outlier_score"), ("velocity", "velocity"),
    ("virality_score", "virality_score"), ("tier", "tier"),
    ("audio_id", "audio_id"), ("audio_title", "audio_title"), ("audio_artist", "audio_artist"),
    ("audio_is_original", "audio_is_original"), ("audio_is_reusable", "audio_is_reusable"),
    ("sound_page_url", "sound_page_url"),
    ("caption", "caption"),
]
RND = {"engagement_rate": 4, "reach_multiplier": 3, "outlier_score": 2, "velocity": 1}


def _cell(r, k):
    v = r.get(k)
    return round(v, RND[k]) if (isinstance(v, float) and k in RND) else v


def _summary(rows):
    by = {}
    for r in rows:
        by.setdefault(r["creator"], []).append(r)
    out = []
    for cr, rs in by.items():
        pl = [r["plays"] for r in rs if r.get("plays")]
        scored = [r for r in rs if r.get("virality_score") is not None]
        top = max(scored, key=lambda r: r["virality_score"]) if scored else None
        out.append({
            "creator": cr, "followers": rs[0].get("creator_followers"),
            "content_analyzed": len(rs),
            "median_plays": int(median(pl)) if pl else None,
            "max_plays": max(pl) if pl else None,
            "viral_count": sum(1 for r in rs if r.get("tier") == "Viral"),
            "top_score": top["virality_score"] if top else None,
            "top_url": top["url"] if top else None,
        })
    out.sort(key=lambda r: (-(r["top_score"] or 0), -(r["followers"] or 0)))
    return out


def write_reports(rows, out_xlsx, out_csv, top_n=100):
    wb = openpyxl.Workbook()
    sh = wb.active; sh.title = "Content"
    sh.append([label for _, label in REEL_COLS])
    for r in rows:
        sh.append([_cell(r, k) for k, _ in REEL_COLS])

    ss = wb.create_sheet("Creator Summary")
    scols = ["creator", "followers", "content_analyzed", "median_plays", "max_plays",
             "viral_count", "top_score", "top_url"]
    ss.append(scols)
    for s in _summary(rows):
        ss.append([s.get(c) for c in scols])

    tv = wb.create_sheet("Top Viral")
    tcols = ["rank", "creator", "url", "tier", "virality_score", "plays", "followers",
             "reach_multiplier", "outlier_score", "engagement_rate", "velocity", "posted", "caption"]
    tv.append(tcols)
    for i, r in enumerate([x for x in rows if x.get("virality_score") is not None][:top_n], 1):
        tv.append([i, r["creator"], r["url"], r.get("tier"), _cell(r, "virality_score"),
                   r.get("plays"), r.get("creator_followers"), _cell(r, "reach_multiplier"),
                   _cell(r, "outlier_score"), _cell(r, "engagement_rate"), _cell(r, "velocity"),
                   r.get("posted_iso"), r.get("caption")])
    Path(out_xlsx).parent.mkdir(parents=True, exist_ok=True)
    # openpyxl streams the workbook into a ZipFile opened on the destination path, so the
    # real file is truncated for the whole serialization. Save to a temp, then rename.
    with atomic_path(out_xlsx) as tmp_xlsx:
        wb.save(tmp_xlsx)

    # The CSV is worse than the xlsx if torn, because a short CSV does NOT raise:
    # core/corpus.py just yields fewer rows, and every producer prompt downstream is then
    # silently built on a partial corpus with no error anywhere. Serialize in memory first
    # so the file on disk is only ever the complete document.
    # Written as BYTES so the line endings are byte-identical to what the old
    # `open(..., newline="")` produced — csv emits \r\n itself, and a text-mode write would
    # translate it again.
    buf = io.StringIO(newline="")
    w = csv.writer(buf)
    w.writerow([label for _, label in REEL_COLS])
    for r in rows:
        w.writerow([_cell(r, k) for k, _ in REEL_COLS])
    write_bytes_atomic(out_csv, buf.getvalue().encode("utf-8"))
